#============================= Languages app =============================#
#                                                                         #
#                   App to make language learning easier                  #
#                                                                         #
#                                                         @FranGarcia94   #
#=========================================================================#


from collections import OrderedDict
from tkinter.ttk import Separator, Style
from typing import Literal
from openpyxl import load_workbook
from fpdf import FPDF
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import easygui as g
import subprocess
from googletrans import Translator
import pyttsx3
import string
from random import *
import re



# Speaker
engine = pyttsx3.init()

# Dark Backgorund
fondo = '#0a1414'
fondo_2 = '#111212'

# Light Background
fondo_3 = '#fcffff'
fondo_4 = '#fcffff'

# Default Background
fondo_5 = '#f0f0f0'


### Interface
root = Tk()
root.iconbitmap("./assets/tr_3.ico")
root.title("Languages App")
root.geometry("975x525")

# Tabs Style
style = Style()
style.theme_use('classic')
style.configure('TNotebook.Tab', background = 'white', foreground = 'blue', font = ('Helvetica', 11))
style.configure('TNotebook', background = fondo_5)
style.configure('TFrame', background = fondo_5, foreground = 'green', font = ('Helvetica', 11))

# Tabs
nb = ttk.Notebook(root)
pesta1 = ttk.Frame(nb)
pesta2 = ttk.Frame(nb)
pesta3 = ttk.Frame(nb)
pesta4 = ttk.Frame(nb)
pesta5 = ttk.Frame(nb, padding=30)

nb.add(pesta1, text = 'Main')
nb.add(pesta2, text = 'Basic verbs')
nb.add(pesta3, text = 'Basic vocabulary')
nb.add(pesta4, text = 'Verbs')
nb.add(pesta5, text = 'Practice')

nb.pack(fill = 'both', expand = 1)

# Frames
mf1 = Frame(pesta1, pady = 10)
mf1.pack()
mf2 = Frame(pesta1)
mf2.pack()

mf6 = LabelFrame(pesta5, text = 'Options', font = ('Tahoma 12 bold'))
mf6.grid(row = 0, column = 0)
mf6.config(padx = 10, pady = 10)

mf7 = Frame(pesta5)
mf7.grid(row = 0, column = 1)

mf8 = Frame(pesta5)
mf8.grid(row = 0, column = 2)


# To add the Scrollbars
def canvas_scrollbar(p):

        main_frame = Frame(p)
        main_frame.pack(fill = BOTH, expand = 1)
        main_frame.config(background = 'blue')
        
        main_frame2 = Frame(p)
        main_frame2.pack(fill = 'both')
        main_frame2.config(background = 'green')

        my_canvas = Canvas(main_frame, bg = fondo_5)
        my_canvas.pack(side = LEFT, fill = 'both', expand = 1)

        my_scrollbar = ttk.Scrollbar(main_frame, orient = VERTICAL, command = my_canvas.yview)
        my_scrollbar.pack(side = RIGHT, fill = Y)

        my_scrollbar_2 = ttk.Scrollbar(main_frame2, orient = HORIZONTAL, command = my_canvas.xview)
        my_scrollbar_2.pack(side = BOTTOM, fill = X)
        
        my_canvas.configure(yscrollcommand = my_scrollbar.set, xscrollcommand = my_scrollbar_2.set)

        my_canvas.bind('<Configure>',lambda e: my_canvas.configure(scrollregion = my_canvas.bbox("all")))

        mf = Frame(my_canvas,bg = fondo_5)

        my_canvas.create_window((0,0), window = mf, anchor = "nw")


        return mf

# Frames with Scrollbar
mf3 = Frame(canvas_scrollbar(pesta2))
mf3.pack()

mf4 = Frame(canvas_scrollbar(pesta3))
mf4.pack()

mf5 = Frame(canvas_scrollbar(pesta4))
mf5.pack()


### Functions ###

def saveExcel():

        workbook = load_workbook(filename = "./lang2.xlsx")

        #open workbook
        sheet = workbook.active = workbook['a']
        
        if esp_Entry.get() == "":

                messagebox.showwarning(title = "WARNING", message = "Field: 'Español' must be filled in")
        else:

                if eng_Entry.get() == "":

                        engVar.set(None)
                if fra_Entry.get() == "":

                        fraVar.set(None)
                if campo_Entry.get() == "":

                        campoVar.set('varios')


                rows = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
                if esp_Entry.get() in rows:

                        op = g.buttonbox(title = 'Found value',
                                msg = '\t\t    That entry already exists. What do you want to do?',
                                choices = ('New meaning', 'Rename', 'Cancel'))

                        if op == 'New meaning':

                                sheet.cell(row = (len(rows)+1), column = 1).value = esp_Entry.get().lower()
                                sheet.cell(row = (len(rows)+1), column = 2).value = eng_Entry.get().lower()
                                sheet.cell(row = (len(rows)+1), column = 3).value = fra_Entry.get().lower()
                                sheet.cell(row = (len(rows)+1), column = 4).value = campo_Entry.get().lower()

                                messagebox.showinfo(title = 'Success',message = 'New meaning added')
                        elif op == 'Rename':

                                pos = rows.index(esp_Entry.get()) + 1 

                                sheet.cell(row = pos, column = 1).value = esp_Entry.get().lower()
                                sheet.cell(row = pos, column = 2).value = eng_Entry.get().lower()
                                sheet.cell(row = pos, column = 3).value = fra_Entry.get().lower()
                                sheet.cell(row = pos, column = 4).value = campo_Entry.get().lower()

                                messagebox.showinfo(title = 'Success', message = 'Renamed value\n'
                                        'Note: Only the first matching element is renamed.')
                        elif op == 'Cancel':

                                pass
                else:

                        sheet.cell(row=(len(rows)+1), column = 1).value = esp_Entry.get().lower()
                        sheet.cell(row=(len(rows)+1), column = 2).value = eng_Entry.get().lower()
                        sheet.cell(row=(len(rows)+1), column = 3).value = fra_Entry.get().lower()
                        sheet.cell(row=(len(rows)+1), column = 4).value = campo_Entry.get().lower()

                        messagebox.showinfo(title = 'Success', message = 'Value added')                        


        workbook.save(filename = "./lang2.xlsx")


def createPDF():

        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook['a']

        rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
        rows2 = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
        rows3 = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]

        listup = []

        for i in range(len(rows1)):

                listup.append((rows1[i], rows2[i], rows3[i]))

        listup.sort(key = lambda x: x[0]) # Sort tuples
        listup.remove(('ESP', 'ENG', 'FRA'))

        #### PDF

        class PDF(FPDF):

                def header(self):

                        self.set_font('Arial', 'B', 10)

                        # Move to the right
                        self.cell(5)

                        # Framed title
                        self.cell(30, 5, 'Language pdf', 1, 0, 'C')

                        # Line break
                        self.ln(10)

                def footer(self):

                        self.set_y(-15)
                        page_num = self.page_no()

                        if page_num > 1 or page_num == 1:

                                self.cell(0,10, str(page_num-1), align = 'R')

        pdf = PDF(orientation = 'P', unit = 'mm', format = 'A4') 
        pdf.add_page()

        # Title
        pdf.set_font('Arial', 'B', 17)
        pdf.set_text_color(0,0,0)

        pdf.cell(w = 0, h = 15, txt = 'Idiomas', border = 1, ln = 1, align = 'C', fill = 0)

        # Header
        pdf.set_font('Arial', 'BU', 15)
        pdf.set_text_color(90, 175, 75)
        
        pdf.cell(w = 63, h = 15, txt = 'Español', border = 1, align = 'C', fill = 0)

        pdf.cell(w = 63, h = 15, txt = 'English', border = 1, align = 'C', fill = 0)

        pdf.multi_cell(w = 0, h = 15, txt = 'Français', border = 1, align = 'C', fill = 0)

        # Values
        pdf.set_font('Arial', '', 13)
        pdf.set_text_color(0, 0, 0)

        for valor in listup:

                # Sheet Frame
                pdf.set_draw_color(0, 0, 255)

                pdf.set_line_width(0.0)
                pdf.line(5.0, 5.0, 205.0, 5.0) # top one
                pdf.line(5.0, 292.0, 205.0, 292.0) # bottom one
                pdf.line(5.0, 5.0, 5.0, 292.0) # left one
                pdf.line(205.0, 5.0, 205.0, 292.0) # right one

                # Set values

                pdf.cell(w = 63, h = 9, txt = str(valor[0]), border = 1, align = 'C', fill = 0)

                pdf.cell(w = 63, h = 9, txt = str(valor[1]), border = 1, align = 'C', fill = 0)

                pdf.multi_cell(w = 0, h = 9, txt = str(valor[2]), border = 1, align = 'C', fill = 0)
        
        pdf.output("./Idiomas_.pdf")

        messagebox.showinfo(title = 'Success', message = 'PDF Created.\nFile name: Idiomas_.pdf')

        workbook.save(filename = "./lang2.xlsx")


def searchWord():

        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook['a']

        rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
        rows2 = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
        rows3 = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]
        rows4 = [str(sheet['D'][i].value) for i in range(len(sheet['A']))]

        def match(entry_get,row):

                if entry_get in row:

                        matching_word = [entry_get]
                else:

                        matching_word = [x for x in row if re.search(f'\\b{entry_get}\\b', x)]
                        print(matching_word)
                
                return matching_word


        if esp_Entry.get() != '' and eng_Entry.get() == '' and fra_Entry.get() == '':

                matching_word = match(esp_Entry.get(), rows1)

                if matching_word != [] or esp_Entry.get() in rows1:
                        
                        espVar.set(rows1[rows1.index(matching_word[0])])
                        engVar.set(rows2[rows1.index(matching_word[0])])
                        fraVar.set(rows3[rows1.index(matching_word[0])])
                        campoVar.set(rows4[rows1.index(matching_word[0])])
                else:

                        messagebox.showwarning(message = 'There are no matches')
        elif eng_Entry.get() != '' and esp_Entry.get() == '' and fra_Entry.get() == '':

                matching_word = match(eng_Entry.get(), rows2)

                if matching_word != [] or eng_Entry.get() in rows2:
                        
                        espVar.set(rows1[rows2.index(matching_word[0])])
                        engVar.set(rows2[rows2.index(matching_word[0])])
                        fraVar.set(rows3[rows2.index(matching_word[0])])
                        campoVar.set(rows4[rows2.index(matching_word[0])])
                else:

                        messagebox.showwarning(message = 'There are no matches')
        elif fra_Entry.get() != '' and esp_Entry.get() == '' and eng_Entry.get() == '':

                matching_word = match(fra_Entry.get(), rows3)

                if matching_word != [] or fra_Entry.get() in rows3:
                        
                        espVar.set(rows1[rows3.index(matching_word[0])])
                        engVar.set(rows2[rows3.index(matching_word[0])])
                        fraVar.set(rows3[rows3.index(matching_word[0])])
                        campoVar.set(rows4[rows3.index(matching_word[0])])
                else:

                        messagebox.showwarning(message = 'There are no matches')
        elif fra_Entry.get() == '' and esp_Entry.get() == '' and eng_Entry.get() == '':

                messagebox.showerror(title = 'Error',message='Blank fields')
        else:

                messagebox.showinfo(message = 'To search, insert a value in any of the language fields, but only in one of them.')
        
        workbook.save(filename = "./lang2.xlsx")


def traductor():

        tr = Translator()

        if esp_Entry.get() != '' and eng_Entry.get() == '' and fra_Entry.get() == '':

                engVar.set(tr.translate(esp_Entry.get(), dest = 'en').text)
                fraVar.set(tr.translate(esp_Entry.get(), dest = 'fr').text)
        elif eng_Entry.get() != '' and esp_Entry.get() == '' and fra_Entry.get() == '':

                espVar.set(tr.translate(eng_Entry.get(), dest = 'es').text)
                fraVar.set(tr.translate(eng_Entry.get(), dest = 'fr').text)
        elif fra_Entry.get() != '' and esp_Entry.get() == '' and eng_Entry.get() == '':

                espVar.set(tr.translate(fra_Entry.get(), dest = 'es').text)
                engVar.set(tr.translate(fra_Entry.get(), dest = 'en').text)
        elif fra_Entry.get() == '' and esp_Entry.get() == '' and eng_Entry.get() == '':

                messagebox.showerror(title = 'Error', message = 'Blanks fields')
        else:

                messagebox.showinfo(message = 'Insert a value in any of the fields, but only one of them')


def voiceFun(lang, i):

        a = []

        if lang != '':

                for voice in engine.getProperty('voices'):

                        a.append(voice.id)

                engine.setProperty('voice', a[i])
                engine.setProperty('rate', 140)
                engine.say(lang)
                engine.runAndWait()  


def reviewWord():

        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook['a']

        t = 1

        rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
        rows2 = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
        rows3 = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]
        
        for i in range(len(rows1)):

                espVar.set(rows1[i])
                engVar.set(rows2[i])
                fraVar.set(rows3[i])
                
                root.update_idletasks()

                voiceFun(esp_Entry.get(), 0)

                root.after(t*300)

                voiceFun(fra_Entry.get(), 3)

        workbook.save(filename = "./lang2.xlsx")        


def color_mode(a,b,c):

        root.config(background = a)
        mf1.config(background = a)
        mf2.config(background = a)

        bb.config(background = b)
        cc.config(background = b)
        dd.config(background = b)
        ee.config(background = b)
        gg.config(background = b)
        delete_button.config(background = b)

        esp_Label.config(background = a,foreground = c)
        eng_Label.config(background = a,foreground = c)
        fra_Label.config(background = a,foreground = c)
        campo_Label.config(background = a,foreground = c)

        style.configure('TFrame', background = a, foreground = 'green', font = ('Helvetica', 11))
        
        esp_Entry.config(highlightbackground = "blue", highlightcolor = "blue")
        eng_Entry.config(highlightbackground = "blue", highlightcolor = "blue")
        fra_Entry.config(highlightbackground = "blue", highlightcolor = "blue")
        campo_Entry.config(highlightbackground = "blue", highlightcolor = "blue")

        if a == fondo_5:

                esp_Entry.config(highlightbackground = a, highlightcolor = a)
                eng_Entry.config(highlightbackground = a, highlightcolor = a)
                fra_Entry.config(highlightbackground = a, highlightcolor = a)
                campo_Entry.config(highlightbackground = a, highlightcolor = a)


### Top Menu Functions ###
def exitApp():

	valor = messagebox.askokcancel("Exit", "Want to get out?")
	if valor == True:
		root.destroy()

def aditionalInfo():

	messagebox.showinfo("Lang.App","Application developed with the purpose of learning languages, in this case, Python.")

def licenseInfo():

	messagebox.showinfo("License","GNU General Public License v3.0")

def openPDF():

        path = "./Idiomas_.pdf"
        subprocess.Popen([path], shell = True)

def openExcel():
        path = './lang2.xlsx'
        subprocess.Popen([path], shell = True)

def deleteFields():

        espVar.set('')
        engVar.set('')
        fraVar.set('')
        campoVar.set('')


def shortcutsEvent(event): # Ctrl + 'letter'

        if event.keysym == 'j':

                deleteFields()
        elif event.keysym == 'p':

                openPDF()
        elif event.keysym == 'e':

                openExcel()
        elif event.keysym == 'l':

                color_mode(fondo_3,fondo_4,fondo)
        elif event.keysym == 'o':

                color_mode(fondo,fondo_2,fondo_3)
        elif event.keysym == 'd':

                color_mode(fondo_5,fondo_5,fondo)
        elif event.keysym == 'b':

                dinamicSearch()


def setSpeaker(arg1, arg2, idioma2):

        a = []
        for voice in engine.getProperty('voices'):

                a.append(voice.id)

        t = 1

        for i in range(len(arg1)):

                engine.setProperty('voice', a[0])
                engine.setProperty('rate', 170)
                engine.say(arg1[i])
                engine.runAndWait()

                root.after(t*100)

                engine.setProperty('voice', a[idioma2])
                engine.setProperty('rate', 170)
                engine.say(arg2[i])
                engine.runAndWait() 


# Tab n
def tabConfig(frame, row1, row2, col1, col2):

        a = []
        b = []

        for i in range(len(row1)):

                a.append(str(row1[i]))
                a[i] = Label(frame, text = str(row1[i]))
                a[i].grid(row = i + 2, column = col1, padx = 3)

                b.append(str(row2[i]))
                b[i] = Label(frame, text = str(row2[i]))
                b[i].grid(row =  i+ 2, column = col2, padx = 3)

                if i == 0 or row1[i] == 'Estaciones' or row1[i-1] == '--------------':

                        a[i].config(font = ('bold', 13), highlightthickness = 2, highlightbackground = 'green')
                        b[i].config(font = ('bold', 13), highlightthickness = 2, highlightbackground = 'green')
                elif row1[i] == 'Presente Simple' or row1[i]=='Pretérito Perfecto' or row1[i] == 'Pluscuamperfecto' or row1[i] == 'Futuro Simple':

                        a[i].config(font = ('bold', 10), highlightthickness = 2, highlightbackground = 'yellow')
                        b[i].config(font = ('bold', 10), highlightthickness = 1, highlightbackground = 'yellow')                       

        button_num = Button(frame, text = "Rep", command = lambda: setSpeaker(row1, row2, 3), image = voice_image)
        button_num.grid(row = 1, column = col1, columnspan = 2)
        button_num.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)

        sep_basico=Separator(frame, orient = 'vertical')
        sep_basico.grid(row = 2, column = col2+1, rowspan = len(row1), sticky = 'ns', padx = 2)


def enter_tab(frame, text_header: string, name_sheet: string):

        head = Label(frame, text = text_header, font = 'bold', bg = 'white', highlightthickness = 4, highlightbackground = 'blue', width = 50)
        head.grid(row = 0, column = 0, columnspan = 11, pady = 10)

        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook[name_sheet]

        numero_columnas = len(sheet[1])
        abc = (string.ascii_uppercase)*10 

        rows = []
        aux = []

        i = 0
        while i < numero_columnas:

                j = 0
                while j <= len(sheet[abc[0]]):

                        try:

                                aux.append(str(sheet[abc[i]][j].value))
                        except:

                                pass

                        j = j+1


                aux = [aux[i] for i in range(len(aux)) if aux[i] != 'None' or i == 0]
                rows.append(aux)
                aux = []
                i = i + 1

        workbook.save(filename = "./lang2.xlsx")

        cont = 0

        for i in range(0, len(rows), 2):

                tabConfig(frame,rows[i], rows[i+1], cont, cont+1)
                cont = cont + 3


def dinamicSearch():

        lb_1 = []
        lb_2 = []
        
        def combos(event):

                if combo2.get() == 'English' or combo2.get() == 'Français':

                        def selection_changed(event):

                                selection = combo.get()
                                rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
                                rows2 = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
                                rows3 = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]
                                rows4 = [str(sheet['D'][i].value) for i in range(len(sheet['A']))]

                                if combo2.get() == 'English':

                                        rows5 = rows2
                                        idioma = 1
                                elif combo2.get() == 'Français':

                                        rows5 = rows3
                                        idioma = 3
                                else:

                                        pass

                                try:

                                        for i in lb_1:

                                                i.destroy()
                                except:
                                        pass

                                try:
                                        for i in lb_2:
                                                i.destroy()
                                except:

                                        pass

                                aux = []
                                aux2 = []
                                aux3 = []
                                
                                for i in range(len(rows1)):

                                        if rows4[i] == selection:

                                                aux3.append((rows1[i], rows5[i]))
                                                
                                aux3.sort(key = lambda x: x[0])

                                for i in range(len(aux3)):

                                        aux.append(aux3[i][0])
                                        aux2.append(aux3[i][1]) 

                                for i in range(len(aux)):

                                        lb_1.append('label'+str(i))
                                        lb_1[i] = Label(mf_1, text = aux[i]) 
                                        lb_1[i].grid(row = i+1, column = 0)

                                        lb_2.append('labe'+str(i))
                                        lb_2[i] = Label(mf_1, text = aux2[i], padx = 10) 
                                        lb_2[i].grid(row = i+1, column = 2)

                                button_=Button(mf_1, text = "Voice", command = lambda: setSpeaker(aux,aux2,idioma))
                                button_.grid(row = 0, column = 0, columnspan = 3)
                                button_.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)
                
                        label_bd = Label(mf_0,text = 'Select the field')
                        label_bd.grid(row = 1, column = 0, columnspan = 2)
                        combo = ttk.Combobox(mf_0, values = rows4_)
                        combo.grid(row = 2, column = 0, columnspan = 2)
                        
                        combo.bind("<<ComboboxSelected>>", selection_changed)


        root2 = Tk()
        root2.iconbitmap("./assets/tr_3.ico")
        root2.title("Dynamic Search")
        root2.geometry("400x400")

        mf_0 = Frame(root2)
        mf_0.pack()
        mf_1 = Frame(canvas_scrollbar(root2))
        mf_1.pack()
        

        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook['a']

        rows4 = [str(sheet['D'][i].value) for i in range(len(sheet['D']))]
        rows4.remove('Campo')
        rows4_ = sorted(list(OrderedDict.fromkeys(rows4)))


        label_lg = Label(mf_0, text = 'Select the field')
        label_lg.grid(row = 0, column = 0, columnspan = 2)
        combo2 = ttk.Combobox(mf_0, values = ['English', 'Français'])
        combo2.grid(row = 0, column = 1, columnspan = 2)
        combo2.set(value = '-Select language-')
        combo2.bind("<<ComboboxSelected>>", combos)
        

        workbook.save(filename =  "./lang2.xlsx")
        root2.mainloop()


def delete_fun():
        
        workbook = load_workbook(filename="./lang2.xlsx")

        sheet = workbook.active=workbook['a']

        rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
        rows2 = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
        rows3 = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]

        try:
                if eng_Entry.get() ==  rows2[rows1.index(esp_Entry.get())] and fra_Entry.get() == rows3[rows1.index(esp_Entry.get())]:

                        delete_index = rows1.index(esp_Entry.get())
                elif esp_Entry.get() ==  rows1[rows2.index(eng_Entry.get())] and fra_Entry.get() == rows3[rows2.index(eng_Entry.get())]:

                        delete_index = rows2.index(eng_Entry.get())
                elif esp_Entry.get() ==  rows1[rows3.index(fra_Entry.get())] and eng_Entry.get() == rows2[rows3.index(fra_Entry.get())]:

                        delete_index = rows3.index(fra_Entry.get())


                ask_delete = messagebox.askquestion('Delete Window', 'Are you sure?')

                if ask_delete == 'yes':

                        sheet.delete_rows(delete_index + 1, 1) # +1 porque los índices en excel empiezan en 1, no en 0.
                        espVar.set('')
                        engVar.set('')
                        fraVar.set('')
                        campoVar.set('')
                        
                        messagebox.showinfo('Delete Window', f'Words deleted: {rows1[delete_index]}, {rows2[delete_index]}, {rows3[delete_index]}')
                else:
                        
                        pass
        except:

                messagebox.showerror('Error', 'Something went wrong')


        workbook.save(filename = "./lang2.xlsx")


# MAIN #

# Tab 1 -#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#

#--------------------VARIABLES--------------------#
espVar = StringVar()
engVar = StringVar()
fraVar = StringVar()
campoVar = StringVar()
varOpcion = IntVar()
#--------------------VARIABLES--------------------#

#--------------------LABEL--------------------#
esp_Label = Label(mf1, text = 'Español', font = 'bold')
esp_Label.grid(row = 0, column = 1, padx = 10, pady = 10)

eng_Label = Label(mf1, text = 'English', font = 'bold')
eng_Label.grid(row = 1, column = 1, padx = 10, pady = 10)

fra_Label = Label(mf1, text = 'Français', font = 'bold')
fra_Label.grid(row = 2, column = 1, padx = 10, pady = 10)

campo_Label = Label(mf1, text = 'Field')
campo_Label.grid(row = 3, column = 1, padx = 10, pady = 5)
#--------------------LABEL--------------------#

#--------------------SEPARATOR--------------------#

def separador(frame, orientation: Literal['horizontal', 'vertical'], r: int = None, rs: int = None, c: int = None, cs: int = None, px: int = None, py: int = None):

        if orientation == 'horizontal':

                st = 'ew'
        else:

                st = 'ns'

        sep_ = Separator(frame, orient = orientation)
        sep_.grid(row = r, rowspan = rs, column = c, columnspan = cs, sticky = st, padx = px, pady = py)


separador(mf1, 'horizontal', 4, None, None, 5, None, 10)

sep_Vmf2 = Separator(mf2, orient = 'vertical')
sep_Vmf2.grid(row = 0, column = 1,rowspan = 3, sticky = 'ns', padx = 30)

sep2_Vmf2 = Separator(mf2, orient = 'vertical')
sep2_Vmf2.grid(row = 0, column = 3,rowspan = 3, sticky = 'ns', padx = 30)

sep2_Vmf2 = Separator(mf2, orient = 'horizontal')
sep2_Vmf2.grid(row = 1, columnspan = 5, sticky = 'ew', pady = 20)
#--------------------SEPARATOR--------------------#

#--------------------ENTRY--------------------#
def check_s(entry_widget):

        def bind_entry(event):# limita los caracteres que puedo escribir

                if len(entry_widget.get()) > 20:

                        #entry_widget['state'] = 'disabled'
                        entry_widget.delete(len(entry_widget.get()) - 1)

        entry_widget.bind("<KeyRelease>",bind_entry)
        

esp_Entry = Entry(mf1, textvariable = espVar, highlightthickness = 2, )
esp_Entry.grid(row = 0, column = 2, padx = 10, pady = 10)

eng_Entry = Entry(mf1, textvariable = engVar, highlightthickness = 2)
eng_Entry.grid(row = 1, column = 2, padx = 10, pady = 10)

fra_Entry = Entry(mf1, textvariable = fraVar, highlightthickness = 2)
fra_Entry.grid(row = 2, column = 2, padx = 10, pady = 10)

campo_Entry = Entry(mf1, textvariable = campoVar, highlightthickness = 2)
campo_Entry.grid(row = 3, column = 2, padx = 10, pady = 5)

entry_list = [esp_Entry, eng_Entry, fra_Entry, campo_Entry]

"""for i in entry_list:
        check_s(i)"""

_ = [check_s(i) for i in entry_list]

#--------------------ENTRY--------------------#

#--------------------BUTTON--------------------#
mf2.config(borderwidth = 5, highlightthickness = 5, highlightbackground = 'lightblue', padx = 5, pady = 10)

bb = Button(mf2, text = "Save", font = 'bold', command = saveExcel)
bb.grid(row = 0, column = 0)
bb.config(padx = 10, pady = 10,width = 10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

cc = Button(mf2, text = "Create PDF", font = 'bold', command = createPDF)
cc.grid(row = 2, column = 0)
cc.config(padx = 10, pady = 10,width = 10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

dd = Button(mf2, text = "Search", font = 'bold', command = searchWord)
dd.grid(row = 0, column = 2)
dd.config(padx = 10, pady = 10,width=10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

ee = Button(mf2, text = "Suggest Translation", font = 'bold', command = traductor)
ee.grid(row = 2, column = 2)
ee.config(padx = 10, pady = 10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

gg = Button(mf2, text = "Review", font = 'bold', command = reviewWord)
gg.grid(row = 2, column = 4)
gg.config(padx = 10, pady = 10,width = 10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

delete_button = Button(mf2, text = "Delete", font = 'bold', command = delete_fun)
delete_button.grid(row = 0, column = 4)
delete_button.config(padx = 10, pady = 10, width = 10, foreground = "darkblue", activeforeground = "#FFA500", bd = 6, cursor = 'hand2')

button_list = [bb, cc, dd, ee, gg]

def bind_fun(dd2):

        def enter_fun(e):

                dd2.config(bg = 'lightgreen')

        def leave_fun(e):

                dd2.config(bg = fondo_5)
                
        dd2.bind('<Enter>', enter_fun)
        dd2.bind('<Leave>', leave_fun)

btn_list = [bb, cc, dd, ee, gg, delete_button]

for btn in btn_list:
        bind_fun(btn)

# Voice Buttons
voice_image = PhotoImage(file = r"./alt_3_2.png")

# Spanish
ff = Button(mf1, text = "ES", command = lambda: voiceFun(esp_Entry.get(), 0), image = voice_image)
ff.grid(row = 0, column = 3)
ff.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)

# English (US)
ff2 = Button(mf1, text = "US", command = lambda: voiceFun(eng_Entry.get(), 1), image = voice_image, compound = 'right')
ff2.grid(row = 1, column = 3)
ff2.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)

# English (UK)
ff2_2 = Button(mf1, text = "UK", command = lambda: voiceFun(eng_Entry.get(), 2), image = voice_image, compound = 'right')
ff2_2.grid(row = 1, column = 4)
ff2_2.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)

# French
ff3 = Button(mf1, text = "FR", command = lambda: voiceFun(fra_Entry.get(), 3), image = voice_image)
ff3.grid(row = 2, column = 3)
ff3.config(padx = 3, pady = 3, foreground = "#2202d9", activeforeground = "#FFA500", bd = 5)
#--------------------BUTTON--------------------#

#--------------------MENU--------------------#
barraMenu = Menu(root)
root.config(menu = barraMenu, width = 350, height = 300)

archivoMenu = Menu(barraMenu, tearoff = 0)
archivoMenu.add_command(label = "Open Excel  Ctrl+E", command = openExcel)
archivoMenu.add_command(label = "Open PDF    Ctrl+P", command = openPDF)
archivoMenu.add_separator()
archivoMenu.add_command(label = "Exit", command = exitApp)

archivoEdicion = Menu(barraMenu, tearoff = 0)
archivoEdicion.add_command(label = "Clear fields           Ctrl+J", command = deleteFields)
archivoEdicion.add_command(label = "Dinamic Search  Ctrl+B", command = dinamicSearch)

archivoHerramientas = Menu(barraMenu,tearoff = 0)
archivoHerramientas.add_command(label = 'Light mode      Ctrl+L', command = lambda: color_mode(fondo_3, fondo_4, fondo))
archivoHerramientas.add_command(label = 'Dark mode   Ctrl+O', command = lambda: color_mode(fondo, fondo_2, fondo_3))
archivoHerramientas.add_command(label = 'Default mode       Ctrl+D', command = lambda: color_mode(fondo_5, fondo_5, fondo))

archivoAyuda = Menu(barraMenu, tearoff = 0)
archivoAyuda.add_command(label = "License", command = licenseInfo)
archivoAyuda.add_command(label = "About", command = aditionalInfo)

barraMenu.add_cascade(label = "File", menu = archivoMenu)
barraMenu.add_cascade(label = "Edition", menu = archivoEdicion)
barraMenu.add_cascade(label = "Tools", menu = archivoHerramientas)
barraMenu.add_cascade(label = "Help", menu = archivoAyuda)
#--------------------MENU--------------------#

#--------------------SHORTCUTS--------------------#
root.bind('<Control-Key-j>', shortcutsEvent) # Clear
root.bind('<Control-Key-p>', shortcutsEvent) # Open PDF
root.bind('<Control-Key-e>', shortcutsEvent) # Open Excel

root.bind('<Control-Key-l>', shortcutsEvent) # Light Mode
root.bind('<Control-Key-o>', shortcutsEvent) # Dark Mode
root.bind('<Control-Key-d>', shortcutsEvent) # Default Mode

root.bind('<Control-Key-b>', shortcutsEvent) # Dinamic Search
#--------------------SHORTCUTS--------------------#


# Tabs -#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#
# Data required: Frame, Header, Excel sheet name

enter_tab(mf3, 'Basic Verbs', 'b')
enter_tab(mf4, 'Basic Vocabulary', 'c')
enter_tab(mf5, 'Verbs', 'd')

# Practice Tab (Game) -#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#-#

modeVar = IntVar()
langVar = IntVar()

font_mf6 = ('Tahoma 10 bold')
font_mf7 = ('Tahoma 12 bold')
font_textbox = ('Consolas 16 bold')

easy = Radiobutton(mf6, text = 'Easy', variable = modeVar, value = 0, takefocus = 0, font = font_mf6)
easy.grid(row = 0, column = 0, padx = 15)
hard = Radiobutton(mf6, text = 'Hard', variable = modeVar, value = 1, takefocus = 0, font = font_mf6)
hard.grid(row = 1, column = 0)

separador(mf6, 'horizontal', 2, None, 0, 2, None, 10)

engmode = Radiobutton(mf6, text = 'English', variable = langVar, value = 0, takefocus = 0, font = font_mf6)
engmode.grid(row = 3, column = 0, padx = 15)
framode = Radiobutton(mf6, text = 'Français', variable = langVar, value = 1, takefocus = 0, font = font_mf6)
framode.grid(row = 4, column = 0)

def playAgain():

        opVar = StringVar()
        workbook = load_workbook(filename = "./lang2.xlsx")
        sheet = workbook.active = workbook['a']

        rows1 = [str(sheet['A'][i].value) for i in range(len(sheet['A']))]
        row_en = [str(sheet['B'][i].value) for i in range(len(sheet['A']))]
        row_fr = [str(sheet['C'][i].value) for i in range(len(sheet['A']))]
        
        if langVar.get() == 0:

                rows3 = row_en
        elif langVar.get() == 1:

                rows3 = row_fr
        else:

                rows3 = rows1

        lst = randint(0, len(rows1))

        ltr_usadas = ['', ' ']
        used_letters = Label(mf7, text = 'Used letters:', font = font_mf7)
        used_letters.grid(row = 6, column = 0, pady = 10)
        used_letters.config(width = 30)

        lives_lbl = Label(mf7, text = 'Lives: 6', font = font_mf7)
        lives_lbl.grid(row = 7, column = 0)
        lives_lbl.config(width = 40)

        hangText = Text(mf8, width = 30, height = 16, highlightthickness = 2, highlightbackground = 'blue', font = font_textbox)
        hangText.grid(row = 10, column = 0, padx = 10, pady = 10)
        hangText.insert(INSERT,
                '       ________________     \n'
                '      |   _____________|    \n'
                '      |  |          |       \n'
                '      |  |         _|_      \n'
                '      |  |        /. .\     \n'
                '      |  |        \_-_/     \n'
                '      |  |          |       \n'
                '      |  |         /|\      \n'
                '      |  |        / | \     \n'
                '      |  |       / / \ \    \n'
                '      |  |      o /   \ o   \n'
                '      |  |       /     \    \n'
                '      |  |      0       0   \n'
                '  ____|__|____              \n'
                ' |____________|             \n')

        lives = [6]
        x = []

        def fun():

                aux = 1
                cont = 0

                for i in rows3[lst]:

                        if i == opVar.get():

                                x[cont] = i 
                                aux = 0

                        cont = cont + 1 
                        
                        lbl = Label(mf7, text = 'Word: ' + " ".join(x), font = font_mf7)
                        lbl.grid(row = 3, column = 0)
                        lbl.config(width = 40)

                if opVar.get() in ltr_usadas or opVar.get() == rows3[lst]:

                        aux = 0
                else:

                        ltr_usadas.append(opVar.get())
                
                used_letters = Label(mf7, text = 'Used: ' + " ".join(ltr_usadas), font = font_mf7)
                used_letters.grid(row = 6, column = 0, pady = 10)
                used_letters.config(width = 40)

                lives[0] = lives[0] - aux
                lives_lbl = Label(mf7, text = 'Lives: ' + str(lives[0]), font = font_mf7)
                lives_lbl.grid(row = 7, column = 0)
                lives_lbl.config(width = 40)

                hangText = Text(mf8, width = 30, height = 16, font = font_textbox)
                hangText.grid(row = 10, column = 0, padx = 10, pady = 10)

                if rows3[lst] == "".join(x) or opVar.get() == rows3[lst]:
                        
                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         /        \n'
                                '      |  |  YEAH            \n'
                                '      |  |         _ _   o  \n'
                                '      |  |        /^ ^\ /   \n'
                                '      |  |        \_V_//    \n'
                                '      |  |         _|_/     \n'
                                '      |  |  ((    / |     ))\n'
                                '      |  |       /  \       \n'
                                '      |  |      /  / \      \n'
                                '      |  |     o  /   \     \n'
                                '  ____|__|____   /    /     \n'
                                ' |____________| 0    0      \n')
                elif lives[0] == 6:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         _|_      \n'
                                '      |  |        /. .\     \n'
                                '      |  |        \_-_/     \n'
                                '      |  |          |       \n'
                                '      |  |         /|\      \n'
                                '      |  |        / | \     \n'
                                '      |  |       / / \ \    \n'
                                '      |  |      o /   \ o   \n'
                                '      |  |       /     \    \n'
                                '      |  |      0       0   \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 5:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |  WTF    _|_      \n'
                                '      |  |        /o o\     \n'
                                '      |  |        \_-_/     \n'
                                '      |  |          |       \n'
                                '      |  |          |\      \n'
                                '      |  |          | \     \n'
                                '      |  |         / \ \    \n'
                                '      |  |        /   \ o   \n'
                                '      |  |       /     \    \n'
                                '      |  |      0       0   \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 4:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         _|_      \n'
                                '      |  |        /o o\     \n'
                                '      |  |        \_~_/     \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |         / \      \n'
                                '      |  |        /   \     \n'
                                '      |  |       /     \    \n'
                                '      |  |      0       0   \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 3:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         _|_      \n'
                                '      |  |        /O O\     \n'
                                '      |  |        \_._/     \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |           \      \n'
                                '      |  |            \     \n'
                                '      |  |             \    \n'
                                '      |  |              0   \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 2:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         _|_      \n'
                                '      |  |     (( /O O\ ))  \n'
                                '      |  |        \_o_/     \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |          |       \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 1:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  | REALLY? _|_      \n'
                                '      |  |        /¬ ¬\     \n'
                                '      |  |        \_-_/     \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '      |  |                  \n'
                                '  ____|__|____              \n'
                                ' |____________|             \n')
                elif lives[0] == 0:

                        hangText.insert(INSERT,
                                '       ________________     \n'
                                '      |   _____________|    \n'
                                '      |  |          |       \n'
                                '      |  |         _|_      \n'
                                '      |  |        /x x\     \n'
                                '      |  |        \_-_/     \n'
                                '      |  |          \       \n'
                                '      |  |         /|\      \n'
                                '      |  |        / | \     \n'
                                '      |  |       | / \ |    \n'
                                '      |  |       o | | o    \n'
                                '      |  |         | |      \n'
                                '      |  |         0 0      \n'
                                '  ____|__|____              \n'
                                ' |____________|  - RIP -    \n')
                

                if rows3[lst] == "".join(x) or opVar.get() == rows3[lst]:

                        messagebox.showinfo(title = 'You WIN', message = f'Congratulations\nThe word is: {rows3[lst]}')
                        playAgain()

                if lives[0] == 0:

                        messagebox.showinfo(title = 'You LOSE', message = f'Try it again\nThe word was: {rows3[lst]}')
                        playAgain()

                opVar.set('')

        for i in rows3[lst]:

                if i == ' ':

                        x.append(' ')
                else:

                        x.append('_')
        

        if modeVar.get() == 0:

                lbl_ = Label(mf7, text = 'Word: ' + rows1[lst], font = font_mf7)
                lbl_.grid(row = 2, column = 0)
                lbl_.config(width = 40)
        else:

                lbl_ = Label(mf7, text = '', font = font_mf7)
                lbl_.grid(row = 2, column = 0)
                lbl_.config(width = 40)

        lbl = Label(mf7, text = 'Word: ' + " ".join(x), font = font_mf7)
        lbl.grid(row = 3, column = 0)
        lbl.config(width = 40)

        ent = Entry(mf7, textvariable = opVar, font = font_mf7)
        ent.grid(row = 4, column = 0, pady = 15)
        ent.config(width = 20, justify = 'center')

        btn = Button(mf7, text = 'Send', command = fun, takefocus = 0, font = font_mf7, bd = 5, bg = 'lightgreen', foreground = 'darkblue', activeforeground = 'lightgreen', activebackground = 'darkblue', cursor='hand2')
        btn.grid(row = 5, column = 0, pady = 10)

        btn2 = Button(mf7, text = 'Play again', command = playAgain, font = font_mf7, bd = 4, bg = 'darkorange',foreground = 'darkblue', activeforeground = 'darkorange', activebackground = 'darkblue', cursor='hand2')
        btn2.grid(row = 8, column = 0, pady = 35)

        workbook.save(filename="./lang2.xlsx")


playAgain()

root.mainloop()